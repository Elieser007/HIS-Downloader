import os
import time

from datetime import datetime

from variables import (
    ESTABLECIMIENTOS,
    DIR_REGISTRO_DIARIO,
    DIR_PLANTILLA_REGISTRO_DIARIO,
    REGISTRO_DIARIO_BASE_SHEET_NAME,
    downloaded_dir_registro_diario_with_start_end,
)

import tkinter as tk
from ttkwidgets import CheckboxTreeview, Calendar
from ttkwidgets.autocomplete import AutocompleteCombobox
from playwright.sync_api import Playwright
from openpyxl import load_workbook

from modules.generics import divide_range_in_days, login


def get_form_registro_diario(page):
    page.frame_locator('frame[name="menuFrame"]').get_by_text("Reportes").click()
    page.frame_locator('frame[name="menuFrame"]').get_by_role(
        "link", name="▾ Registro Diario de Consultas"
    ).click()


def registro_diario_downloader(
    playwright: Playwright, startDate, endDate, seleccionados, diag_new, unify_base
):
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://hisguaira.mspbs.gov.py/ambulatoria/")
    login(page)
    get_form_registro_diario(page)
    date_range = divide_range_in_days(startDate, endDate)
    SAVE_AS_DOWNLOAD = downloaded_dir_registro_diario_with_start_end(startDate, endDate)

    for selected in seleccionados:
        page.frame_locator('frame[name="mainFrame"]').locator(
            '//*[@id="form1"]/div/div[1]/div/span/span[1]/span/span[2]'
        ).click()
        time.sleep(1)
        page.frame_locator('frame[name="mainFrame"]').get_by_role(
            "option", name=selected, exact=True
        ).click()
        time.sleep(1)

        for date in date_range:
            page.frame_locator('frame[name="mainFrame"]').locator(
                'input[name="startDate"]'
            ).fill(date[0])
            page.frame_locator('frame[name="mainFrame"]').locator(
                'input[name="endDate"]'
            ).fill(date[1])
            if diag_new:
                page.frame_locator('frame[name="mainFrame"]').get_by_label("Si").nth(
                    1
                ).check()
            time.sleep(1)
            page.frame_locator('frame[name="mainFrame"]').get_by_role(
                "button", name="Generar"
            ).click()
            time.sleep(1)

            try:
                with page.expect_download() as download_info:
                    page.frame_locator('frame[name="mainFrame"]').locator(
                        "#form1"
                    ).get_by_role("img").click(timeout=300000)
                download = download_info.value
            except:
                login(page)
                get_form_registro_diario(page)
                page.frame_locator('frame[name="mainFrame"]').locator(
                    'input[name="startDate"]'
                ).fill(date[0])
                page.frame_locator('frame[name="mainFrame"]').locator(
                    'input[name="endDate"]'
                ).fill(date[1])
                if diag_new:
                    page.frame_locator('frame[name="mainFrame"]').get_by_label(
                        "Si"
                    ).nth(1).check()
                page.frame_locator('frame[name="mainFrame"]').get_by_role(
                    "button", name="Generar"
                ).click()
                time.sleep(1)
                with page.expect_download() as download_info:
                    page.frame_locator('frame[name="mainFrame"]').locator(
                        "#form1"
                    ).get_by_role("img").click(timeout=300000)
                download = download_info.value

            download.save_as(SAVE_AS_DOWNLOAD + download.suggested_filename)
            time.sleep(3)

    context.close()
    browser.close()
    if unify_base:
        unify_base_registro_diario(SAVE_AS_DOWNLOAD)


def form_download_registro_diario(playwright):
    root = tk.Tk()
    root.geometry("500x768")
    global diag_new
    diag_new = False
    global unify_base
    unify_base = True
    tree_label = tk.Label(root, text="Seleccione Establecimientos")
    tree_label.pack()
    tree = CheckboxTreeview(root)
    tree.insert("", "end", "todos", text="Seleccionar Todos")
    for est in ESTABLECIMIENTOS:
        tree.insert(
            "todos",
            "end",
            est,
            text=est,
        )
    tree.pack()

    def def_diag_new():
        global diag_new
        diag_new = not diag_new

    def def_unify_base():
        global unify_base
        unify_base = not unify_base

    v_d_new = tk.BooleanVar()
    d_new = tk.Checkbutton(
        root,
        text="¿Diagnostico Nuevo?",
        variable=v_d_new,
        onvalue=True,
        offvalue=False,
        command=def_diag_new,
    )
    d_new.pack()

    v_u_base = tk.BooleanVar()
    u_base = tk.Checkbutton(
        root,
        text="¿Unificar y crear Base después de descargar?",
        variable=v_u_base,
        onvalue=True,
        offvalue=False,
        command=def_unify_base,
    )
    u_base.pack()
    u_base.select()

    start_label = tk.Label(root, text="Fecha de Inicio")
    start_label.pack()
    start_date = Calendar(
        root,
        selectforeground="white",
        selectbackground="red",
        locale="es_ES",
    )
    start_date.pack()

    end_label = tk.Label(root, text="Fecha de Termino")
    end_label.pack()
    end_date = Calendar(
        root,
        selectforeground="white",
        selectbackground="red",
        locale="es_ES",
    )
    end_date.pack()

    tk.Button(
        root,
        pady=2,
        text="Descargar Registro Diario",
        command=lambda: root.quit(),
    ).pack(pady=5)
    root.mainloop()
    seleccionados = tree.get_checked()
    start_selected = str(start_date.selection).split(" ")[0]
    end_selected = str(end_date.selection).split(" ")[0]
    registro_diario_downloader(
        playwright, start_selected, end_selected, seleccionados, diag_new, unify_base
    )
    root.quit()


def form_crear_base_registro_diario():
    root = tk.Tk()
    root.geometry("400x200")
    path_dir = os.scandir(DIR_REGISTRO_DIARIO)
    list_values = []
    for path in path_dir:
        list_values.append(path.name)
    combo_label = tk.Label(root, text="Seleccione Carpeta de Archivos Descargados")
    combo_label.pack()
    combo_box = AutocompleteCombobox(root, width=60, completevalues=list_values)
    combo_box.pack()
    tk.Button(
        root,
        pady=2,
        text="Crear Base de Registro Diario",
        command=lambda: root.quit(),
    ).pack(pady=5)
    root.mainloop()
    unify_base_registro_diario(combo_box.get())
    root.quit()


def unify_base_registro_diario(folder_selected):

    wb_base = load_workbook(DIR_PLANTILLA_REGISTRO_DIARIO)
    ws_base = wb_base[REGISTRO_DIARIO_BASE_SHEET_NAME]

    fila_insercion_base = 3

    ultimo_insertado = fila_insercion_base

    INFORME_REGISTRO_DIARIO_SHEET_NAME = "reporte_registro_diario_consult"
    fila_copia_informe = 7

    column_number_format = [8, 11]
    column_date_format=[1]

    files = os.scandir(os.path.join(DIR_REGISTRO_DIARIO, folder_selected))
    desktop = os.path.join(os.path.join(os.environ["USERPROFILE"]), "Desktop")
    for file in files:
        wb = load_workbook(file.path)
        ws = wb[INFORME_REGISTRO_DIARIO_SHEET_NAME]

        filas = ws[f"A{fila_copia_informe}":f"BX{ws.max_row}"]

        for fila in filas:
            for celda in fila:
                if celda.column in column_number_format:
                    num_int = 0
                    try:
                        num_int = int(celda.value)
                    except ValueError:
                        num_int = celda.value
                    ws_base.cell(row=ultimo_insertado, column=celda.column).value = (
                        num_int
                    )
                elif(celda.column in column_date_format):
                    date_format=None
                    try:
                        date_format = datetime.strptime(celda.value, '%Y-%m-%d').date()
                    except ValueError:
                        date_format=celda.value
                    ws_base.cell(row=ultimo_insertado, column=celda.column).value = (
                        date_format
                    )
                else:
                    ws_base.cell(row=ultimo_insertado, column=celda.column).value = (
                        celda.value
                    )

            ultimo_insertado = ultimo_insertado + 1

        wb.close()
    wb_base.save(
        os.path.join(
            desktop,
            f"Registro_Diario_{str(datetime.now().strftime('%Y-%m-%d_%H.%M.%S.hs.xlsx'))}",
        )
    )
    wb_base.close()
